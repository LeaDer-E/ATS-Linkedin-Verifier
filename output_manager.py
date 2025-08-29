import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter


def display_results_terminal(results):
    print("=" * 60)
    for res in results:
        status_icon = "✔" if res["status"] else "✖"
        print(f"📄 File: {res['file_name']}")
        print("=" * 60)
        print(f"📍 Path: {res['full_path']}")
        print(f"📁 Folder: {res['folder_name']}")
        print(f"📞 Phone: {res['phone']}")
        print(f"🔗 Link: {res['link']}")
        print(f"👤 Name: {res['name'] if res['name'] else 'N/A'}")
        print(f"✅ Status: {status_icon}")
        print("-" * 60)

def save_results_html(results, output_folder):
    """Save results to a dark-themed HTML file with a status filter."""
    os.makedirs(output_folder, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_path = os.path.join(output_folder, f"linkedin_results_{timestamp}.html")

    html_content = """
    <html>
    <head>
    <meta charset="UTF-8">
    <title>LinkedIn Check Results</title>
    <style>
        body { background-color: #121212; color: white; font-family: Arial, sans-serif; }
        table { border-collapse: collapse; width: 100%; margin-top: 20px; }
        th, td { border: 1px solid #444; padding: 8px; text-align: left; }
        th { background-color: #1f1f1f; }
        tr:nth-child(even) { background-color: #1e1e1e; }
        .status-true { color: #4CAF50; font-weight: bold; }
        .status-false { color: #F44336; font-weight: bold; }
        select { margin-top: 10px; padding: 5px; }
        a { color: #4da3ff; text-decoration: none; }
        a:hover { text-decoration: underline; }
    </style>
    <script>
        function filterStatus() {
            var filter = document.getElementById("statusFilter").value;
            var rows = document.getElementById("resultsTable").rows;
            for (var i = 1; i < rows.length; i++) {
                var statusCell = rows[i].cells[6];
                if (filter === "all") {
                    rows[i].style.display = "";
                } else if (filter === "valid" && statusCell.innerText === "✔") {
                    rows[i].style.display = "";
                } else if (filter === "invalid" && statusCell.innerText === "✖") {
                    rows[i].style.display = "";
                } else {
                    rows[i].style.display = "none";
                }
            }
        }
    </script>
    </head>
    <body>
    <h1>LinkedIn Check Results</h1>
    <label for="statusFilter">Filter by status: </label>
    <select id="statusFilter" onchange="filterStatus()">
        <option value="all">All</option>
        <option value="valid">✔ Valid</option>
        <option value="invalid">✖ Invalid</option>
    </select>
    <table id="resultsTable">
    <tr>
        <th>File Name</th>
        <th>Phone</th>
        <th>Full Path</th>
        <th>Link</th>
        <th>Folder Name</th>
        <th>Name</th>
        <th>Status</th>
    </tr>
    """

    for res in results:
        status_icon = "✔" if res["status"] else "✖"
        status_class = "status-true" if res["status"] else "status-false"
        link_html = f'<a href="{res["link"]}" target="_blank">{res["link"]}</a>'
        
        # استخراج مجلد الملف من full_path
        folder_path = os.path.dirname(res['full_path']).replace('\\', '/')
        # الرابط يفتح المجلد فقط
        folder_link = f'<a href="file:///{folder_path}" target="_blank">{res["full_path"]}</a>'
        
        html_content += f"""
        <tr>
            <td>{res['file_name']}</td>
            <td>{res['phone']}</td>
            <td>{folder_link}</td>
            <td>{link_html}</td>
            <td>{res['folder_name']}</td>
            <td>{res['name'] if res['name'] else 'N/A'}</td>
            <td class="{status_class}">{status_icon}</td>
        </tr>
        """


    html_content += """
    </table>
    </body>
    </html>
    """

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    print(f"[HTML] Results saved to {output_path}")




def save_results_excel(results, output_folder):
    """Save results to an Excel file with formatting."""
    os.makedirs(output_folder, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    output_path = os.path.join(output_folder, f"linkedin_results_{timestamp}.xlsx")

    df = pd.DataFrame(results)
    df["status"] = df["status"].apply(lambda x: "✔" if x else "✖")
    df = df[["file_name", "phone", "full_path", "link", "folder_name", "name", "status"]]
    df.columns = ["File Name", "Phone", "Full Path", "Link", "Folder Name", "Name", "Result"]

    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    # Freeze first column (A) and first row (1)
    ws.freeze_panes = "B2"

    # Yellow fill for header row A1:G1
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Define full border style (all sides)
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    # Font size 16 for all cells
    font_16 = Font(size=16)

    # Set the height of the first row to 40 (or double if already set)
    ws.row_dimensions[1].height = ws.row_dimensions[1].height * 2 if ws.row_dimensions[1].height else 40

    # Set column widths
    ws.column_dimensions['A'].width = 30.00
    ws.column_dimensions['B'].width = 30.00
    ws.column_dimensions['C'].width = 50.00
    ws.column_dimensions['D'].width = 50.00
    ws.column_dimensions['E'].width = 35.00
    ws.column_dimensions['F'].width = 35.00
    ws.column_dimensions['G'].width = 10.00

    # Apply yellow fill, font size 16, border, and center alignment to header row only
    for col in range(1, 8):  # Columns A=1 to G=7
        cell = ws.cell(row=1, column=col)
        cell.fill = yellow_fill
        cell.font = Font(size=16, bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # For the rest of the cells (from row 2 down), apply font size 16, full border, and center alignment
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if col == 1:
                cell.font = Font(size=16, bold=True)
            else:
                cell.font = Font(size=16, bold=False)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Link column (3) = Full Path, URL to File Explorer
        cell_fullpath = ws.cell(row=row, column=3)
        full_path = cell_fullpath.value
        if full_path and not str(full_path).startswith('=HYPERLINK'):
            folder_path = os.path.dirname(full_path).replace('\\', '\\\\')
            cell_fullpath.value = f'=HYPERLINK("file:///{folder_path}", "{full_path}")'
            cell_fullpath.font = Font(color="0000FF", underline="single", size=16)

        # Link column (4) = LinkedIn URL
        cell_link = ws.cell(row=row, column=4)
        url = cell_link.value
        if url and not str(url).startswith('=HYPERLINK'):
            cell_link.value = f'=HYPERLINK("{url}", "{url}")'
            cell_link.font = Font(color="0000FF", underline="single", size=16)




    wb.save(output_path)
    print(f"[Excel] Results saved to {output_path}")